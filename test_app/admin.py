# test_app/admin.py
from django.contrib import admin
from django.urls import path
from django.shortcuts import render, redirect
from django.contrib import messages
from django.http import HttpResponseRedirect
import pandas as pd
from .models import (
    Question,
    Test,
    TestSession,
    TestResult,
    StudentAnswer,
    Subject,
    Topic
)

# NEW: Inline admin for Topics under Subject
class TopicInline(admin.TabularInline):
    model = Topic
    extra = 1
    fields = ['name', 'description', 'order']
    show_change_link = True

# NEW: Subject Admin with inline topics
@admin.register(Subject)
class SubjectAdmin(admin.ModelAdmin):
    list_display = ['name', 'class_level', 'created_at']
    list_filter = ['class_level']
    search_fields = ['name', 'description']
    inlines = [TopicInline]
    
    fieldsets = (
        ('Subject Information', {
            'fields': ('name', 'class_level', 'description')
        }),
    )

# NEW: Topic Admin for standalone management
@admin.register(Topic)
class TopicAdmin(admin.ModelAdmin):
    list_display = ['name', 'subject', 'order', 'created_at']
    list_filter = ['subject__class_level', 'subject']
    search_fields = ['name', 'subject__name']
    list_editable = ['order']
    
    fieldsets = (
        ('Topic Information', {
            'fields': ('name', 'subject', 'description', 'order')
        }),
    )

# Updated Question Admin
@admin.register(Question)
class QuestionAdmin(admin.ModelAdmin):
    list_display = ['text', 'question_type', 'get_subject', 'get_topic']
    list_filter = ['question_type']  # Remove the invalid filters for now
    search_fields = ['text']
    
    def get_subject(self, obj):
        if obj.test and obj.test.subject_fk:
            return obj.test.subject_fk.name
        elif obj.test and obj.test.subject:
            return obj.test.subject
        return '-'
    get_subject.short_description = 'Subject'
    
    def get_topic(self, obj):
        if obj.test and obj.test.topic_fk:
            return obj.test.topic_fk.name
        elif obj.test and obj.test.topic:
            return obj.test.topic
        return '-'
    get_topic.short_description = 'Topic'
    
    # Add custom URL for Excel upload
    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path('upload-questions/', self.upload_questions, name='upload_questions'),
        ]
        return custom_urls + urls
    
    def changelist_view(self, request, extra_context=None):
        from django.contrib import messages
        from django.utils.safestring import mark_safe
        
        extra_context = extra_context or {}
        extra_context['upload_url'] = 'upload-questions/'
        
        # Add a button message
        messages.info(request, mark_safe(
            '<a href="upload-questions/" style="background-color: #28a745; color: white; padding: 8px 15px; text-decoration: none; border-radius: 4px; display: inline-block;">'
            '📤 Click here to upload questions from Excel'
            '</a>'
        ))
        
        return super().changelist_view(request, extra_context=extra_context)
    
    
    def upload_questions(self, request):
        if request.method == 'POST':
            excel_file = request.FILES.get('excel_file')
            
            if not excel_file:
                messages.error(request, 'Please select an Excel file')
                return HttpResponseRedirect('../')
            
            # Check file extension
            file_ext = excel_file.name.split('.')[-1].lower()
            if file_ext not in ['xlsx', 'xls']:
                messages.error(request, 'Please upload a valid Excel file (.xlsx or .xls)')
                return HttpResponseRedirect('../')
            
            try:
                import openpyxl
                from io import BytesIO
                
                # Read the Excel file
                file_data = BytesIO(excel_file.read())
                
                # Try to load the workbook
                try:
                    workbook = openpyxl.load_workbook(file_data, data_only=True)
                except Exception as e:
                    messages.error(request, f'Could not read Excel file: {str(e)}')
                    return HttpResponseRedirect('../')
                
                # Get the active sheet
                sheet = workbook.active
                if sheet is None:
                    messages.error(request, 'Excel file has no worksheets')
                    return HttpResponseRedirect('../')
                
                # Check if sheet has any rows
                if sheet.max_row is None or sheet.max_row < 1:
                    messages.error(request, 'Excel file is empty')
                    return HttpResponseRedirect('../')
                
                # Get headers from first row (handle None values)
                headers = []
                for col in range(1, sheet.max_column + 1):
                    if sheet.max_column is None:
                        break
                    cell_value = sheet.cell(row=1, column=col).value
                    if cell_value:
                        headers.append(str(cell_value).strip().lower())
                
                if not headers:
                    messages.error(request, 'Could not read headers from Excel file. Please ensure the first row contains column names.')
                    return HttpResponseRedirect('../')
                
                # Required columns (case insensitive)
                required_columns = ['subject_name', 'topic_name', 'class_level', 'question_text', 'question_type', 
                                'option_a', 'option_b', 'option_c', 'option_d', 'correct_answer']
                
                # Map actual headers to required columns (case insensitive)
                header_map = {}
                for req_col in required_columns:
                    for header in headers:
                        if header.lower() == req_col.lower():
                            header_map[req_col] = header
                            break
                
                # Check for missing columns
                missing_columns = [col for col in required_columns if col not in header_map]
                if missing_columns:
                    messages.error(request, f'Missing columns: {", ".join(missing_columns)}. Found headers: {headers}')
                    return HttpResponseRedirect('../')
                
                # Get column indices
                col_indices = {}
                for col_num, header in enumerate(headers, 1):
                    for req_col in required_columns:
                        if header.lower() == req_col.lower():
                            col_indices[req_col] = col_num
                            break
                
                created_count = 0
                error_count = 0
                errors = []
                
                # Process each row (starting from row 2)
                for row_num in range(2, sheet.max_row + 1):
                    try:
                        # Get values from each column (handle None values)
                        subject_name = sheet.cell(row=row_num, column=col_indices['subject_name']).value
                        topic_name = sheet.cell(row=row_num, column=col_indices['topic_name']).value
                        class_level = sheet.cell(row=row_num, column=col_indices['class_level']).value
                        question_text = sheet.cell(row=row_num, column=col_indices['question_text']).value
                        question_type = sheet.cell(row=row_num, column=col_indices['question_type']).value
                        option_a = sheet.cell(row=row_num, column=col_indices['option_a']).value
                        option_b = sheet.cell(row=row_num, column=col_indices['option_b']).value
                        option_c = sheet.cell(row=row_num, column=col_indices['option_c']).value
                        option_d = sheet.cell(row=row_num, column=col_indices['option_d']).value
                        correct_answer = sheet.cell(row=row_num, column=col_indices['correct_answer']).value
                        
                        # Skip empty rows
                        if not question_text or str(question_text).strip() == '':
                            continue
                        
                        # Convert to string and strip
                        subject_name = str(subject_name).strip() if subject_name else ''
                        topic_name = str(topic_name).strip() if topic_name else ''
                        class_level = str(class_level).strip().upper() if class_level else ''
                        question_text = str(question_text).strip()
                        
                        # Validate required fields
                        if not subject_name:
                            errors.append(f"Row {row_num}: Missing subject_name")
                            error_count += 1
                            continue
                        if not topic_name:
                            errors.append(f"Row {row_num}: Missing topic_name")
                            error_count += 1
                            continue
                        if not class_level:
                            errors.append(f"Row {row_num}: Missing class_level")
                            error_count += 1
                            continue
                        
                        # Process question_type
                        if not question_type:
                            question_type = 'mcq'
                        else:
                            question_type = str(question_type).lower().strip()
                            if question_type not in ['mcq', 'theory']:
                                question_type = 'mcq'
                        
                        # Get or create Subject
                        subject, _ = Subject.objects.get_or_create(
                            name=subject_name,
                            class_level=class_level
                        )
                        
                        # Get or create Topic
                        topic, _ = Topic.objects.get_or_create(
                            name=topic_name,
                            subject=subject
                        )
                        
                        # Get or create Test
                        test, created = Test.objects.get_or_create(
                            class_level=class_level,
                            subject_fk=subject,
                            topic_fk=topic,
                            defaults={
                                'created_by': request.user,
                                'subject': subject_name,
                                'topic': topic_name
                            }
                        )
                        
                        if not created and not test.created_by:
                            test.created_by = request.user
                            test.save()
                        
                        # Get option values
                        opt_a = str(option_a).strip() if option_a else ''
                        opt_b = str(option_b).strip() if option_b else ''
                        opt_c = str(option_c).strip() if option_c else ''
                        opt_d = str(option_d).strip() if option_d else ''
                        correct = str(correct_answer).strip().upper() if correct_answer else ''
                        
                        # Create Question
                        Question.objects.create(
                            test=test,
                            text=question_text,
                            question_type=question_type,
                            option_a=opt_a,
                            option_b=opt_b,
                            option_c=opt_c,
                            option_d=opt_d,
                            correct_answer=correct
                        )
                        
                        created_count += 1
                        
                    except Exception as e:
                        error_count += 1
                        errors.append(f"Row {row_num}: {str(e)}")
                
                # Show results
                if created_count > 0:
                    messages.success(request, f'✅ Successfully uploaded {created_count} questions!')
                
                if error_count > 0:
                    messages.warning(request, f'⚠️ {error_count} errors. First few: {"; ".join(errors[:5])}')
                
                if created_count == 0 and error_count == 0:
                    messages.warning(request, 'No questions found in the Excel file. Please check your data.')
                
            except Exception as e:
                import traceback
                traceback.print_exc()
                messages.error(request, f'Error reading Excel file: {str(e)}')
            
            return HttpResponseRedirect('../')
        
        return render(request, 'admin/upload_questions.html', {
            'title': 'Upload Questions from Excel',
            'opts': self.model._meta,
        })
        

@admin.register(Test)
class TestAdmin(admin.ModelAdmin):
    list_display = ['get_subject_name', 'get_topic_name', 'class_level', 'created_by', 'date_created']
    search_fields = ['subject_fk__name', 'topic_fk__name', 'subject', 'topic']
    list_filter = ['class_level', 'date_created']
    
    def get_subject_name(self, obj):
        if obj.subject_fk:
            return obj.subject_fk.name
        return obj.subject
    get_subject_name.short_description = 'Subject'
    
    def get_topic_name(self, obj):
        if obj.topic_fk:
            return obj.topic_fk.name
        return obj.topic
    get_topic_name.short_description = 'Topic'
    
    fieldsets = (
        ('Test Information', {
            'fields': ('class_level', 'subject_fk', 'topic_fk', 'subject', 'topic')
        }),
    )
    
    # Auto-set created_by to current user
    def save_model(self, request, obj, form, change):
        if not change:  # If creating new object
            obj.created_by = request.user
        super().save_model(request, obj, form, change)
    
    # Filter form fields
    def formfield_for_foreignkey(self, db_field, request, **kwargs):
        if db_field.name == "topic_fk":
            kwargs["queryset"] = Topic.objects.select_related('subject').all()
        return super().formfield_for_foreignkey(db_field, request, **kwargs)
    

# Keep existing admin registrations for other models
@admin.register(TestSession)
class TestSessionAdmin(admin.ModelAdmin):
    list_display = ['id', 'student', 'class_level', 'subject', 'topic', 'created_at', 'completed']
    list_filter = ['class_level', 'subject', 'completed']
    search_fields = ['student__username', 'subject', 'topic']

@admin.register(StudentAnswer)
class StudentAnswerAdmin(admin.ModelAdmin):
    list_display = ['result', 'question', 'selected_option', 'is_correct', 'theory_match_percentage']
    list_filter = ['is_correct']
    search_fields = ['result__student__username', 'question__text']

class StudentAnswerInline(admin.TabularInline):
    model = StudentAnswer
    extra = 0
    readonly_fields = ['question', 'selected_option', 'is_correct', 'theory_answer', 'theory_match_percentage']

@admin.register(TestResult)
class TestResultAdmin(admin.ModelAdmin):
    list_display = ['student', 'test_session', 'score', 'created_at']
    search_fields = ['student__username', 'test_session__subject']
    list_filter = ['created_at']
    inlines = [StudentAnswerInline]